"""
Utils for common I/O operations, including datasets and input/output from
command prompt.
"""
import csv
import sys

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import os.path
from pandas import DataFrame, Series
import pandas as pd
import xlsxwriter
import xlwings


#####################################################################
class IOUtilsError(Exception):
    """Encapsulates known errors in this module.

    Attributes:
        method (str): Name of method in which error was encountered.
    """

    def __init__ (self, method, msg=None):
        if msg is None:
            msg = 'Error encountered in {}.'.format(method)
        Exception.__init__(self, msg)
        self.method = method


#####################################################################

class PrintUpdate(object):

    def __init__ (self):
        self.max_update_length = 0

    def print (self, msg):
        """Handles writing updates to stdout when we want to clear previous
        outputs to the console (or other stdout).
        """
        self.max_update_length = max(self.max_update_length, len(msg))
        empty_chars = self.max_update_length - len(msg)
        msg = '{0}{1}'.format(msg, ' '*empty_chars)
        sys.stdout.write('{}\r'.format(msg))
        sys.stdout.flush()


#####################################################################
def response_yes (question):
    """Returns true if user types Y or y in command prompt to given question."""
    question = '\n' + question + ' Enter y or n:\n'
    response = input(question).strip().lower()
    if response=='y':
        return True
    elif response=='n':
        return False

    print('\nInvalid response supplied before so trying again.\n')
    return response_yes(question)


def response_no (question):
    return not response_yes(question)


def section_break (divider=None, lines=0):
    if divider is not None:
        output = divider
    else:
        output = '-'*40
    if lines>0:
        output += ('\n'*lines)
    print(output)


#####################################################################
def get_all_dir_files (folder_path):
    """Returns names of all files located in `folder_path`.

    Notes:
        Results will not include files in any subdirectories within
        `folder_path`.
        Only file names (not full file paths) are returned.
    """
    return [f for f in os.listdir(folder_path)
            if os.path.isfile(os.path.join(folder_path, f))]


#####################################################################
def save_df_to_excel (df, path, wks='Main', cols=None, idx=True,
                      force_dir=False):
    """Saves DataFrame to Excel file.

    Args:
        df (DataFrame): Data to save.
        path (str): File path to save DataFrame to.
        wks (str): Optional. Name of worksheet data saved to.
        cols (sequence): Optional. Columns to write.
        idx (bool): Default True. Write row names (index).
        force_dir (bool): Default False. If True, then a folder is made to
        accommodate the implied folder from the file path. If False, an error
        is raised if such folder doesn't already exist.
    """
    # If df_entries is a Series, clean it to a DataFrame.
    if isinstance(df, Series):
        df = df.to_frame()
    path = clean_excel_path(path)
    # Check whether directory exists.
    directory = os.path.dirname(path)
    if not os.path.isdir(directory):
        if not force_dir:
            raise FileNotFoundError("Directory doesn't exist: {}".format(
                  directory))
        os.makedirs(directory)
    # Create Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(path, engine='xlsxwriter')
    # Convert DataFrame to XlsxWriter Excel object.
    df.to_excel(writer, sheet_name=wks, columns=cols, index=idx)
    # Close the Pandas Excel writer / output the Excel file.
    writer.save()


def save_dfs_to_excel (dfs, path, wks_names=None, cols=None, idx=None,
                       force_dir=False):
    """Saves multiple dfs to Excel workbook under different worksheets.

    Args:
        dfs (list): DataFrames to save.
        path (str): File path.
        wks_names (list): Optional. Names of worksheets where dfs saved. If
        not provided, each worksheet will be named "Main_i" where i is the
        index of the sheet (starting at 0).
        cols (list): Optional. Sequences containing columns to write for each
        df_entries.
        idx (list): Optional. Booleans for whether to write row labels for
        given df_entries. Default is to write row labels for all dfs.
        force_dir (bool): Default False. If True, then directory is made if
        that implied by path does not exist.
    """
    # If any of the dfs is a Series, clean it to DataFrame.
    for i, df in enumerate(dfs):
        if isinstance(df, pd.Series):
            dfs[i] = dfs[i].to_frame()
    path = clean_excel_path(path)
    # Check whether directory exists.
    directory = os.path.dirname(path)
    if not os.path.isdir(directory):
        if not force_dir:
            raise FileNotFoundError("Directory doesn't exist: {}".format(
                  directory))
        os.makedirs(directory)
    # Init worksheet names if none provided.
    if wks_names is None:
        wks_names = []
        for i, _ in enumerate(dfs):
            wks_names.append('Main_{}'.format(i))
    # Create Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(path, engine='xlsxwriter')

    for i, df in enumerate(dfs):
        # Determine whether to only write specific cols.
        if cols is None:
            cols = None
        else:
            cols = cols[i]
        # Determine whether to write row names (idx).
        if idx is None:
            write_idx = True
        else:
            write_idx = idx[i]
        df.to_excel(writer, sheet_name=wks_names[i], columns=cols,
                    index=write_idx)
    # Close the Pandas Excel writer / output the Excel file.
    writer.save()


def save_data_to_excel (data, path, wks_name='Main'):
    """Writes data to Excel file."""
    # Create a workbook and add worksheet.
    path = clean_excel_path(path)
    wkb = xlsxwriter.Workbook(path)
    wks = wkb.add_worksheet(wks_name)

    # Output each list as a row in worksheet.
    for row, values in enumerate(data):
        for col, val in enumerate(values):
            wks.write(row, col, val)

    wkb.close()


def save_data_to_csv (data, path, delimiter=',', lineterminator='\n'):
    """Writes data to CSV file.

    Args:
        data (list[list]): Matrix of data to write.
        path (str): Save file path.
    """
    path = clean_csv_path(path)
    with open(path, 'w') as f:
        wtr = csv.writer(f, delimiter=delimiter, lineterminator=lineterminator)
        wtr.writerows(data)


def write_list_to_excel (data, path, wks_name='Main'):
    """Writes list to Excel file in column A and saves to path.

    Args:
        data (list): What to write.
        path (str): File path.
        wks_name (str): Name of sheet which will contain data.
    """
    # Ensure data is a sequence.
    try:
        test_iterator = iter(data)
    except TypeError:
        raise IOUtilsError('write_list_to_excel', 'data is not an iterable.')
    else:
        if len(data)==0:
            raise IOUtilsError('write_list_to_excel', 'data is empty.')

    # Create a workbook and add worksheet.
    path = clean_excel_path(path)
    wkb = xlsxwriter.Workbook(path)
    wks = wkb.add_worksheet(wks_name)

    # Output each item from list on new row in worksheet.
    for row, value in enumerate(data):
        wks.write(row, 0, value)

    wkb.close()


#####################################################################
def read_excel_df_xlwings (path, wks_name, anchor_cell):
    """Reads Excel worksheet data into DataFrame using xlwings.

    Notes:
        The index of the DataFrame is automatically set to the first column
        by xlwings.

    Args:
        path (str): File path.
        wks_name (str): Name of worksheet containing data.
        anchor_cell (tuple): (row, column) for cell from which to expand (
        1-index based).

    Returns:
          DataFrame
    """
    row, col = anchor_cell
    app = xlwings.App(visible=False)  # must be initiated to keep Excel hidden
    wkb = xlwings.Book(path)
    try:
        wks = wkb.sheets(wks_name)
        cell = wks.cells(row, col)
        df = wks.range(cell1=cell).expand().options(pd.DataFrame).value
    finally:
        wkb.close()
    return df


def read_excel_data_openpyxl (path, wks_name, first_row=1, last_row=None,
                              first_col=1, last_col=None):
    """Reads Excel data using openpyxl lib.
    
    Notes:
        openpyxl row and column index begins at 1.
    
    Args:
        path (str): File path.
        wks_name (str): Name of worksheet containing data.
        first_row (int): First row containing data.
        last_row (int): Optional. Last row containing data of interest. If
        not provided, inferred from the worksheet.
        first_col (int): First column containing data.
        last_col (int): Optional. Last column containing data. If not
        provided, inferred from the worksheet.
        
    Returns:
          data (list): Each element is another list containing the column
          values for that row.
    """
    path = ensure_excel_path_valid(path)
    wkb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    wks = wkb[wks_name]

    if last_row is None:
        last_row = WksTools.get_last_row(wks, first_col)
    if last_col is None:
        last_col = WksTools.get_last_col(wks, first_row)

    data = []
    for row in range(first_row, last_row + 1):
        row_data = []
        for col in range(first_col, last_col + 1):
            value = wks.cell(row=row, column=col).value
            row_data.append(value)
        data.append(row_data)
    return data


class WksTools(object):
    """Static class for utils related to openpyxl Worksheets."""

    @staticmethod
    def get_last_row (wks, col, iter_start=1000):
        """Returns row of first cell encountered in col column with value
        that is not None.

        Args:
            wks (Worksheet): Worksheet we care about.
            col (int): Column to use as proxy for the entire sheet.
            iter_start (int): Row from which to begin iterating upward.

        Iterates upward beginning at cell coordinates (iter_start, col) to find
        cell containing value.
        """
        for row in range(iter_start, 0, -1):
            if wks.cell(row=row, column=col).value is not None:
                return row
        return 1

    @staticmethod
    def get_last_col (wks, row, iter_start=100):
        """Returns column of first cell encountered in row with value that is
        not None.

        Args:
            wks (Worksheet): Worksheet we care about.
            row (int): Row to use as proxy for the entire sheet.
            iter_start (int): Column from which to begin iterating leftward.

        Iterates leftward beginning at cell coordinates (row, iter_start) to
        find cell containing value.
        """
        for col in range(iter_start, 0, -1):
            if wks.cell(row=row, column=col)._value is not None:
                return col
        return 1


#####################################################################
def read_lines_from_txt (path):
    """Reads list of values contained as separate lines in txt file."""
    path = ensure_txt_path_valid(path)
    with open(path, 'r') as f:
        return f.read().splitlines()


#####################################################################
def clean_csv_path (path):
    """Ensures path is a valid CSV file path by ensuring that it (1) contains
    .csv extension and (2) doesn't already exist. If it does, adds an
    appendage to make it a unique.
    """
    # Ensure path contains csv extension.
    if path[-4:]!='.csv':
        path += '.csv'
    # Ensure path doesn't already exist.
    if os.path.exists(path):
        path_without_ext = path[:-4]
        version_count = 1
        while os.path.exists(path):
            path = path_without_ext + str(version_count) + '.csv'
    return path


def clean_excel_path (path):
    """Ensures path is a valid Excel file path by ensuring that it (1)
    contains .xlsx file extension [not the .xls extension] and (2) doesn't
    already exist. If it does, adds an appendage to make it unique.
    """
    # First ensure the Excel file path is properly formatted.
    if _path_contains_xls_ext(path):
        _remove_excel_xls_ext(path)
    if not _path_contains_xlsx_ext(path):
        path += '.xlsx'
    if os.path.exists(path):
        # Append a V[x] to the file path to create a unique path.
        path_without_ext = path[:-5]
        version_count = 1
        while os.path.exists(path):
            path = '{0} V{1}.xlsx'.format(path_without_ext, version_count)
            version_count += 1
    return path


def get_existing_excel_path_from_user (prompt):
    """Requests path to existing Excel file from user, calling itself
    recursively until a valid Excel path is provided.
    """
    if prompt is None:
        prompt = '\nEnter Excel file path:\n'
    else:
        prompt = '\n' + prompt + '\n'
    path = input(prompt)

    # Add .xlsx extension to path if it contains neither .xls or .xlsx.
    if not _path_contains_xlsx_ext(path) and not _path_contains_xls_ext(path):
        path += '.xlsx'
    if os.path.isfile(path):
        return path

    prompt = 'No Excel file at following path: {0}\n. Enter again:'.format(path)
    return get_existing_excel_path_from_user(prompt)


def get_existing_csv_path_from_user (prompt=None):
    """Requests path to existing CSV file from user, calling itself
    recursively until a valid path is provided.
    """
    if prompt is None:
        prompt = '\nEnter CSV file path:\n'
    else:
        prompt = '\n' + prompt + '\n'

    path = input(prompt)

    if not _path_contains_csv_ext(path):
        path += '.csv'
    if os.path.isfile(path):
        return path

    prompt = 'No CSV at following path: {0}\n. Enter again:'.format(path)
    return get_existing_csv_path_from_user(prompt)


def get_new_excel_path_from_user (prompt=None):
    """Requests path to use to save a new Excel file."""
    if prompt is None:
        prompt = '\nEnter Excel file path:\n'
    else:
        prompt = '\n' + prompt + '\n'
    path = input(prompt)

    # Add .xlsx extension to path if it doesn't contain an extension.
    if not _path_contains_an_excel_extension(path):
        path += '.xlsx'
    return path


def ensure_excel_path_valid (path):
    """Returns version of path verified to be full file path to existing
    Excel file.
    """
    if not _path_contains_xlsx_ext(path) and not _path_contains_xls_ext(path):
        # If path contains no Excel file extension, try each one until we
        # find a path that exists.
        new_path = path + '.xlsx'
        if os.path.isfile(new_path):
            return new_path
        else:
            new_path = path + '.xls'
            if os.path.isfile(new_path):
                return new_path
            info = 'No .xls or .xlsx file found at path : {0}'.format(path)
            raise FileNotFoundError(info)
    if os.path.isfile(path):
        # Path contains Excel file extension.
        return path
    info = 'No file found at path: {}'.format(path)
    raise FileNotFoundError(info)


def ensure_csv_path_valid (path):
    """Returns version of path verified to be full file path to existing CSV
    file or raise FileNotFoundError.
    """
    if path[-4:]!='.csv':
        path += '.csv'
    if os.path.isfile(path):
        return path
    raise FileNotFoundError('No csv file found at path: {}'.format(path))


def ensure_txt_path_valid (path):
    """Returns version of path verified to be full file path to existing .txt
    file.
    """
    if path[-4:]!='.txt':
        path += '.txt'
    if os.path.isfile(path):
        return path
    raise FileNotFoundError('No txt file found at path: {0}'.format(path))


def _path_contains_an_excel_extension (path):
    """Returns True if path contains either .xls or .xlsx extension."""
    return _path_contains_xlsx_ext(path) or _path_contains_xls_ext(path)


def _path_contains_xlsx_ext (path):
    """Returns True if path's file extension is equivalent to the .xlsx
    extension.
    """
    return path[-5:]=='.xlsx'


def _path_contains_xls_ext (path):
    """Returns True if path's file extension is equivalent to the .xls
    extension.
    """
    return path[-4:]=='.xls'


def _path_contains_csv_ext (path):
    return path[-4:]=='.csv'


def _remove_excel_xls_ext (path):
    """Removes last 4 characters of path (assuming they're the .xls
    extension).
    """
    return path[:-4]
